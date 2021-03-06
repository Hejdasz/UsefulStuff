import os, re, pprint, openpyxl, sys, shelve
from openpyxl import Workbook

'''
shows all the files in folder
'''

print('start')      
a=input()

path = os.getcwd()
path_num = len(path.split('\\'))

split_path = path.split('\\')
title = split_path[-1]
wb = openpyxl.Workbook()
sheet = wb['Sheet']
row_num = 2
highest_column = 1

for folderName, subfolders, filenames in os.walk(path):
    split_folderName = folderName.split('\\')
    highest_column_help = len(split_folderName)

    if highest_column_help > highest_column:
        highest_column = highest_column_help
        
sheet.cell(row=1, column=1).value = 'Folder'
for a in range (2,highest_column):
    sheet.cell(row=1, column=a).value = 'SubFolder'    
sheet.cell(row=1, column=highest_column).value = 'File'

for folderName, subfolders, filenames in os.walk(path):
    fs = list(filenames)
    fs.sort()
    
    split_folderName = folderName.split('\\')
    
    for f in range (1,len(split_folderName) ):
        sheet.cell(row=row_num, column=f).value = str(split_folderName[f])
        sheet.cell(row=row_num, column=highest_column).value = 'Empty'        
    
    for a in fs:
        b = len(split_folderName)-1
        
        sheet.cell(row=row_num, column=highest_column).value = str(a)

        for d in range (0,b):
        
            sheet.cell(row=row_num, column=b-d).value = str(split_folderName[b-d]) #AAAAAAAAAAAA
                
        row_num+=1


wb.save(filename = title+'.xlsx')
print("\n"+"end")

xd = input()
